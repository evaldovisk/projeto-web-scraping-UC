import openpyxl


class Excel:
    def __init__(self, path_excel, nome_sheet, numero_linha_inicial_dados):
        self.path_excel = path_excel
        self.nome_sheet = nome_sheet
        self.numero_linha_inicial_dados = numero_linha_inicial_dados

        self.arquivo_excel_aberto = openpyxl.load_workbook(self.path_excel)
        self.pagina_seleciona = self.arquivo_excel_aberto[self.nome_sheet]

    # recolhe todos os documentos junto ao digito
    # int, int -> list: int
    def captura_todos_cpf(self, numero_coluna_documento, numero_coluna_digito):
        coluna_documento = numero_coluna_documento
        coluna_digito = numero_coluna_digito

        lista_todos_documentos = list()

        for linha in self.pagina_seleciona.iter_rows(min_row=self.numero_linha_inicial_dados):
            documento = str(linha[coluna_documento].value)
            digito = str(linha[coluna_digito].value)

            if len(digito) == 1:
                digito = '0' + digito

            cpf = documento + digito
            if not cpf == 'NoneNone':
                lista_todos_documentos.append(cpf)

        return lista_todos_documentos

    # Recolhe todos os numeros de contrato
    # int, int -> list: int
    def captura_todos_contratos(self, numero_coluna_contrato):
        coluna_contrato = numero_coluna_contrato
        lista_todos_contractos = list()

        for linha in self.pagina_seleciona.iter_rows(min_row=self.numero_linha_inicial_dados):
            contrato = str(linha[coluna_contrato].value)
            if not contrato == None:
                lista_todos_contractos.append(contrato)

        return lista_todos_contractos

    # Inseri todas as ucs
    # int, list: str -> None
    def salvar_ucs_excel(self, numero_coluna_uc, lista_uc):
        coluna_uc = numero_coluna_uc

        for row, uc in zip(self.pagina_seleciona.iter_rows(min_row=self.numero_linha_inicial_dados), lista_uc):
            row[coluna_uc].value = uc

        nome_salve = self.path_excel[:len(self.path_excel)-5] + '-scrapyng.xlsx'
        self.arquivo_excel_aberto.save(nome_salve)

